from typing import Dict, Any, List
from pathlib import Path
import logging
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
import json

logger = logging.getLogger(__name__)


class ExcelReportGenerator:
    """Generate Excel reports from report data"""
    
    def __init__(self):
        self.header_font = Font(bold=True, color="FFFFFF", size=12)
        self.header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
        self.header_alignment = Alignment(horizontal="center", vertical="center")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    async def generate(
        self, 
        report_data: Dict[str, Any], 
        report_id: str,
        output_dir: Path
    ) -> str:
        """Generate Excel report and return file path"""
        try:
            filename = f"report_{report_id}.xlsx"
            file_path = output_dir / filename
            
            # Create workbook
            wb = Workbook()
            
            # Remove default sheet
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])
            
            # Add summary sheet
            self._create_summary_sheet(wb, report_data)
            
            # Add metrics sheet if available
            if "data" in report_data and "metrics" in report_data["data"]:
                self._create_metrics_sheet(wb, report_data["data"]["metrics"])
            
            # Add trade log sheet if this is a trade log report
            if report_data.get("report_type") == "trade_log" and "data" in report_data:
                self._create_trade_log_sheet(wb, report_data["data"])
            
            # Add grouping sheets
            if "data" in report_data and "groupings" in report_data["data"]:
                for grouping_name, grouping_data in report_data["data"]["groupings"].items():
                    self._create_grouping_sheet(wb, grouping_name, grouping_data)
            
            # Add charts sheet if available
            if "data" in report_data and "charts" in report_data["data"]:
                self._create_charts_sheet(wb, report_data["data"]["charts"])
            
            # Save workbook
            wb.save(str(file_path))
            
            logger.info(f"Excel report generated: {file_path}")
            return str(file_path)
            
        except Exception as e:
            logger.error(f"Error generating Excel report: {str(e)}")
            raise
    
    def _create_summary_sheet(self, wb: Workbook, report_data: Dict[str, Any]):
        """Create summary sheet"""
        ws = wb.create_sheet("Summary")
        
        # Title
        ws["A1"] = f"{report_data.get('report_type', 'Report').replace('_', ' ').title()}"
        ws["A1"].font = Font(bold=True, size=16)
        ws.merge_cells("A1:D1")
        
        # Period
        if "period" in report_data:
            ws["A3"] = "Report Period:"
            ws["B3"] = f"{report_data['period']['start']} to {report_data['period']['end']}"
            ws["A3"].font = Font(bold=True)
        
        # Generated timestamp
        ws["A4"] = "Generated:"
        ws["B4"] = report_data.get("generated_at", datetime.utcnow().isoformat())
        ws["A4"].font = Font(bold=True)
        
        # Summary data
        if "summary" in report_data:
            row = 6
            ws[f"A{row}"] = "Summary Statistics"
            ws[f"A{row}"].font = Font(bold=True, size=14)
            row += 2
            
            # Headers
            ws[f"A{row}"] = "Metric"
            ws[f"B{row}"] = "Value"
            for col in ["A", "B"]:
                ws[f"{col}{row}"].font = self.header_font
                ws[f"{col}{row}"].fill = self.header_fill
                ws[f"{col}{row}"].alignment = self.header_alignment
            
            row += 1
            
            # Data
            for key, value in report_data["summary"].items():
                ws[f"A{row}"] = key.replace("_", " ").title()
                if isinstance(value, float):
                    if "rate" in key:
                        ws[f"B{row}"] = f"{value*100:.1f}%"
                    else:
                        ws[f"B{row}"] = round(value, 2)
                else:
                    ws[f"B{row}"] = value
                row += 1
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
    
    def _create_metrics_sheet(self, wb: Workbook, metrics: Dict[str, Any]):
        """Create metrics sheet"""
        ws = wb.create_sheet("Metrics")
        
        # Title
        ws["A1"] = "Performance Metrics"
        ws["A1"].font = Font(bold=True, size=16)
        ws.merge_cells("A1:D1")
        
        # Metrics table
        row = 3
        ws[f"A{row}"] = "Metric"
        ws[f"B{row}"] = "Value"
        ws[f"C{row}"] = "Description"
        
        for col in ["A", "B", "C"]:
            ws[f"{col}{row}"].font = self.header_font
            ws[f"{col}{row}"].fill = self.header_fill
            ws[f"{col}{row}"].alignment = self.header_alignment
        
        row += 1
        
        # Metric descriptions
        descriptions = {
            "total_pnl": "Total profit/loss across all trades",
            "win_rate": "Percentage of winning trades",
            "avg_trade": "Average profit/loss per trade",
            "sharpe_ratio": "Risk-adjusted return metric",
            "max_drawdown": "Maximum peak-to-trough decline"
        }
        
        for key, value in metrics.items():
            ws[f"A{row}"] = key.replace("_", " ").title()
            
            if isinstance(value, float):
                if "rate" in key or "percent" in key:
                    ws[f"B{row}"] = f"{value*100:.1f}%"
                elif "pnl" in key or "profit" in key or "loss" in key:
                    ws[f"B{row}"] = f"${value:,.2f}"
                else:
                    ws[f"B{row}"] = round(value, 4)
            else:
                ws[f"B{row}"] = value
            
            ws[f"C{row}"] = descriptions.get(key, "")
            row += 1
        
        self._auto_adjust_columns(ws)
    
    def _create_trade_log_sheet(self, wb: Workbook, trades_data: Any):
        """Create trade log sheet"""
        ws = wb.create_sheet("Trade Log")
        
        # Extract trades list
        if isinstance(trades_data, dict) and "data" in trades_data:
            trades = trades_data["data"]
        elif isinstance(trades_data, list):
            trades = trades_data
        else:
            trades = []
        
        if not trades:
            ws["A1"] = "No trades found"
            return
        
        # Headers
        headers = [
            "ID", "Symbol", "Strategy", "Entry Date", "Exit Date",
            "Entry Price", "Exit Price", "Quantity", "P&L", "P&L %",
            "Fees", "Notes"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border
        
        # Data
        for row, trade in enumerate(trades, 2):
            ws.cell(row=row, column=1, value=trade.get("id", ""))
            ws.cell(row=row, column=2, value=trade.get("symbol", ""))
            ws.cell(row=row, column=3, value=trade.get("strategy", ""))
            ws.cell(row=row, column=4, value=trade.get("entry_date", ""))
            ws.cell(row=row, column=5, value=trade.get("exit_date", ""))
            ws.cell(row=row, column=6, value=trade.get("entry_price", 0))
            ws.cell(row=row, column=7, value=trade.get("exit_price", 0))
            ws.cell(row=row, column=8, value=trade.get("quantity", 0))
            
            pnl = trade.get("pnl", 0)
            pnl_cell = ws.cell(row=row, column=9, value=pnl)
            if pnl < 0:
                pnl_cell.font = Font(color="FF0000")  # Red for losses
            else:
                pnl_cell.font = Font(color="008000")  # Green for profits
            
            ws.cell(row=row, column=10, value=trade.get("pnl_percent", 0))
            ws.cell(row=row, column=11, value=trade.get("fees", 0))
            ws.cell(row=row, column=12, value=trade.get("notes", ""))
        
        # Format columns
        ws.column_dimensions["D"].number_format = "yyyy-mm-dd"
        ws.column_dimensions["E"].number_format = "yyyy-mm-dd"
        ws.column_dimensions["F"].number_format = "$#,##0.00"
        ws.column_dimensions["G"].number_format = "$#,##0.00"
        ws.column_dimensions["I"].number_format = "$#,##0.00"
        ws.column_dimensions["J"].number_format = "0.00%"
        ws.column_dimensions["K"].number_format = "$#,##0.00"
        
        # Add filters
        ws.auto_filter.ref = ws.dimensions
        
        self._auto_adjust_columns(ws)
    
    def _create_grouping_sheet(self, wb: Workbook, grouping_name: str, grouping_data: Dict[str, Any]):
        """Create sheet for grouping analysis"""
        sheet_name = grouping_name.replace("_", " ").title()[:31]  # Excel sheet name limit
        ws = wb.create_sheet(sheet_name)
        
        # Title
        ws["A1"] = sheet_name
        ws["A1"].font = Font(bold=True, size=16)
        ws.merge_cells("A1:D1")
        
        # Table
        row = 3
        headers = ["Group", "Trade Count", "Total P&L", "Win Rate", "Avg Trade"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
        
        row += 1
        
        # Data
        for group, stats in grouping_data.items():
            ws.cell(row=row, column=1, value=str(group))
            ws.cell(row=row, column=2, value=stats.get("count", 0))
            ws.cell(row=row, column=3, value=stats.get("pnl", 0))
            ws.cell(row=row, column=4, value=stats.get("win_rate", 0))
            ws.cell(row=row, column=5, value=stats.get("pnl", 0) / stats.get("count", 1) if stats.get("count", 0) > 0 else 0)
            row += 1
        
        # Format columns
        ws.column_dimensions["C"].number_format = "$#,##0.00"
        ws.column_dimensions["D"].number_format = "0.00%"
        ws.column_dimensions["E"].number_format = "$#,##0.00"
        
        # Add chart if there's data
        if row > 4:  # Has data rows
            chart = BarChart()
            chart.title = f"P&L by {sheet_name}"
            chart.y_axis.title = "P&L ($)"
            chart.x_axis.title = "Group"
            
            data = Reference(ws, min_col=3, min_row=3, max_row=row-1, max_col=3)
            categories = Reference(ws, min_col=1, min_row=4, max_row=row-1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)
            
            ws.add_chart(chart, f"G{row + 2}")
        
        self._auto_adjust_columns(ws)
    
    def _create_charts_sheet(self, wb: Workbook, charts_data: Dict[str, Any]):
        """Create charts sheet with data tables"""
        ws = wb.create_sheet("Charts Data")
        
        # Title
        ws["A1"] = "Chart Data"
        ws["A1"].font = Font(bold=True, size=16)
        
        current_row = 3
        
        # Equity curve data
        if "equity_curve" in charts_data and "data" in charts_data["equity_curve"]:
            ws[f"A{current_row}"] = "Equity Curve"
            ws[f"A{current_row}"].font = Font(bold=True, size=14)
            current_row += 1
            
            # Headers
            ws[f"A{current_row}"] = "Date"
            ws[f"B{current_row}"] = "Cumulative P&L"
            ws[f"C{current_row}"] = "Trade P&L"
            
            for col in ["A", "B", "C"]:
                ws[f"{col}{current_row}"].font = self.header_font
                ws[f"{col}{current_row}"].fill = self.header_fill
            
            current_row += 1
            start_row = current_row
            
            # Data
            for point in charts_data["equity_curve"]["data"]:
                ws[f"A{current_row}"] = point["date"]
                ws[f"B{current_row}"] = point["cumulative_pnl"]
                ws[f"C{current_row}"] = point["trade_pnl"]
                current_row += 1
            
            # Add line chart
            if current_row > start_row:
                chart = LineChart()
                chart.title = "Equity Curve"
                chart.y_axis.title = "Cumulative P&L ($)"
                chart.x_axis.title = "Date"
                
                data = Reference(ws, min_col=2, min_row=start_row-1, max_row=current_row-1, max_col=2)
                dates = Reference(ws, min_col=1, min_row=start_row, max_row=current_row-1)
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(dates)
                
                ws.add_chart(chart, f"E{start_row}")
            
            current_row += 3
        
        # Monthly P&L data
        if "monthly_pnl" in charts_data and "data" in charts_data["monthly_pnl"]:
            ws[f"A{current_row}"] = "Monthly P&L"
            ws[f"A{current_row}"].font = Font(bold=True, size=14)
            current_row += 1
            
            # Headers
            ws[f"A{current_row}"] = "Month"
            ws[f"B{current_row}"] = "P&L"
            
            for col in ["A", "B"]:
                ws[f"{col}{current_row}"].font = self.header_font
                ws[f"{col}{current_row}"].fill = self.header_fill
            
            current_row += 1
            start_row = current_row
            
            # Data
            for point in charts_data["monthly_pnl"]["data"]:
                ws[f"A{current_row}"] = point["month"]
                ws[f"B{current_row}"] = point["pnl"]
                current_row += 1
            
            # Add bar chart
            if current_row > start_row:
                chart = BarChart()
                chart.title = "Monthly P&L"
                chart.y_axis.title = "P&L ($)"
                chart.x_axis.title = "Month"
                
                data = Reference(ws, min_col=2, min_row=start_row-1, max_row=current_row-1, max_col=2)
                categories = Reference(ws, min_col=1, min_row=start_row, max_row=current_row-1)
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(categories)
                
                ws.add_chart(chart, f"E{start_row}")
        
        self._auto_adjust_columns(ws)
    
    def _auto_adjust_columns(self, ws):
        """Auto-adjust column widths based on content"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width